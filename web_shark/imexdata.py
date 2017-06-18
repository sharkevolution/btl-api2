# -*- coding: UTF-8 -*-

# Импорт данных из файла єксель

import os
import csv
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment


def dispatcher_extension(dest_filename, NAMEsheet):
    """
    Диспетчер запуска чтения файла от расширения
    :param dest_filename:
    :return:
    """
    tmpfile, fileExtension = os.path.splitext(dest_filename)
    result = []

    if fileExtension == '.txt':
        result = readTextfile(dest_filename)
    elif fileExtension == '.xlsx':
        result = readExcel(dest_filename, NAMEsheet)

    return result


def readTextfile(dest_filename):
    pull_figure = []
    fg = read_string_txt(dest_filename)
    next(fg)
    for b in fg:
        pull_figure.append([int(b[0]), int(b[1])])

    return pull_figure


def read_string_txt(dest_filename):
    """
    Чтение данных из текстового файла с использованием генераторов
    """
    with open(dest_filename, 'r') as csvfile:
        spamreader = csv.reader(csvfile, delimiter='\t',
                                quoting=csv.QUOTE_NONE,
                                lineterminator='\n')
        for row in spamreader:
            yield row


def readExcel(dest_filename, NAMEsheet):
    wb = load_workbook(filename=dest_filename)
    sheet_ranges = wb[NAMEsheet]
    rw = 1

    pull_figure = []

    while sheet_ranges.cell(column=1, row=rw + 1).value:
        frsize = sheet_ranges.cell(column=1, row=rw + 1).value
        frcount = sheet_ranges.cell(column=2, row=rw + 1).value
        rw += 1
        pull_figure.append([frsize, frcount])

    return pull_figure


def saveExcel(resdict, filename):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "alignment"

    name_column = 'ABCDEFGHKLMNOPQR'
    for b in name_column:
        ws1.column_dimensions[b].width = 4

    font = Font(name='Calibri',
                size=9,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False)

    alignment = Alignment(horizontal='center',
                          vertical='center',
                          text_rotation=0,
                          wrap_text=False,
                          shrink_to_fit=False,
                          indent=0)

    font_knife = Font(name='Calibri',
                      size=12,
                      bold=True,
                      italic=False,
                      vertAlign=None,
                      underline='none',
                      strike=False)

    knife = [1, 2, 3, 4, 5, 6, 7, 8, 'balance']
    for col in range(0, 9):
        vl = ws1.cell(column=col + 1, row=1, value=knife[col])
        vl.font = font_knife
        vl.alignment = alignment

    long_strips = resdict[1]
    matrix_result = resdict[0]['matrix_result']
    for row in range(0, len(matrix_result)):
        pull = matrix_result[row]
        for col in range(0, 8):
            if col < len(pull):
                if pull[col] > 0:
                    vl = ws1.cell(column=col + 1, row=row + 2, value=pull[col])
                    vl.font = font
                    vl.alignment = alignment
            else:
                break

        bl = long_strips - sum(pull)
        vl = ws1.cell(column=9, row=row + 2, value=bl)
        if bl > 0:
            vl.font = Font(name='Calibri', size=12, bold=True)
        else:
            vl.font = Font(name='Calibri', size=12, bold=False)
        vl.alignment = alignment

    matrix_state = resdict[0]['matrix_state']
    for row in range(0, len(matrix_state)):
        pull = matrix_state[row]
        for col in range(0, 8):
            if col < len(pull):
                vl = ws1.cell(column=col + 11, row=row + 2, value=pull[col])
                vl.font = font
                vl.alignment = alignment
            else:
                break

    _ = ws1.cell(column=20, row=1, value='full_line')
    _ = ws1.cell(column=21, row=1, value=resdict[0]['full_line'])
    _ = ws1.cell(column=20, row=2, value='bad_line')
    _ = ws1.cell(column=21, row=2, value=resdict[0]['bad_line'])
    _ = ws1.cell(column=20, row=3, value='permutation')
    _ = ws1.cell(column=21, row=3, value=resdict[0]['permutation'])

    wb.save(filename=filename)

    return 0


def export_excel(data, filename):
    """
        Сохранить сырую матрицу в файл эксель
    :param resdict: 
    :param filename: 
    :return: 
    """
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "data"

    name_column = 'AB'
    for b in name_column:
        ws1.column_dimensions[b].width = 10

    font = Font(name='Calibri',
                size=9,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False)

    alignment = Alignment(horizontal='center',
                          vertical='center',
                          text_rotation=0,
                          wrap_text=False,
                          shrink_to_fit=False,
                          indent=0)

    font_title = Font(name='Calibri',
                      size=12,
                      bold=True,
                      italic=False,
                      vertAlign=None,
                      underline='none',
                      strike=False)

    title_export = ['figure', 'count']
    for col in range(0, 2):
        vl = ws1.cell(column=col + 1, row=1, value=title_export[col])
        vl.font = font_title
        vl.alignment = alignment

    row = 2
    for d in data.items():
        m = d[1]

        vl = ws1.cell(column=1, row=row, value=m['f'])
        vl.font = font
        vl.alignment = alignment

        vl = ws1.cell(column=2, row=row, value=m['c'])
        vl.font = font
        vl.alignment = alignment

        row += 1

    wb.save(filename=filename)

    return 0

